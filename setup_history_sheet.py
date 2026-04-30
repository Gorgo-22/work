from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 既存のExcelを開く
wb = load_workbook("food_prices.xlsx")

# 既に「履歴」シートがあれば削除して作り直す
if "履歴" in wb.sheetnames:
    del wb["履歴"]

ws = wb.create_sheet(title="履歴")

# ヘッダー行
ヘッダー = [
    "期間開始日",
    "期間終了日",
    "東京基準値",
    "東京前週比",
    "大阪基準値",
    "大阪前週比",
    "名古屋基準値",
    "名古屋前週比",
    "玉子市況",
    "報告書ファイル名",
    "実行日時",
]
ws.append(ヘッダー)

# ヘッダー行を装飾
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# 列幅
列幅 = {
    "A": 12, "B": 12,
    "C": 12, "D": 12,
    "E": 12, "F": 12,
    "G": 14, "H": 14,
    "I": 22,
    "J": 30,
    "K": 18,
}
for 列, 幅 in 列幅.items():
    ws.column_dimensions[列].width = 幅

# 履歴シートを末尾に配置
wb.save("food_prices.xlsx")
print("「履歴」シートを追加しました！")
print(f"現在のシート構成: {wb.sheetnames}")