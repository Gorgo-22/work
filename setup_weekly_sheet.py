from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 既存のExcelを開く
wb = load_workbook("food_prices.xlsx")

# 既に「週次設定」シートがあれば削除して作り直す
if "週次設定" in wb.sheetnames:
    del wb["週次設定"]

ws = wb.create_sheet(title="週次設定")

# 設定項目を書き込む
設定項目 = [
    ["項目", "値"],
    ["期間開始日", "2026/4/24"],
    ["期間終了日", "2026/4/30"],
    ["東京玉子基準値", 315],
    ["東京前週比", "±0"],
    ["大阪玉子基準値", 315],
    ["大阪前週比", "±0"],
    ["名古屋玉子基準値", 315],
    ["名古屋前週比", "±0"],
    ["玉子市況", "保合（小幅変動横ばい）"],
    ["玉子発表日", "2026/4/24"],
]

for row_data in 設定項目:
    ws.append(row_data)

# ヘッダー行を装飾
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# 列幅
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 30

# シートを先頭に移動（最初に開かれるように）
wb.move_sheet("週次設定", offset=-len(wb.sheetnames) + 1)

wb.save("food_prices.xlsx")
print("「週次設定」シートを追加しました！")
print("Excelで food_prices.xlsx を開いて、「週次設定」シートに毎週の値を入れてください。")