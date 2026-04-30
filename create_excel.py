import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 新しいExcelブックを作成
wb = Workbook()
ws = wb.active
ws.title = "食材価格"

# CSVを読み込んでExcelに書き込む
with open("food_prices.csv", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, start=1):
        for col_idx, value in enumerate(row, start=1):
            # 数値の列(C列とD列、つまり先月・今月)は数値として保存
            if row_idx > 1 and col_idx in (3, 4):
                ws.cell(row=row_idx, column=col_idx, value=int(value))
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

# ヘッダー行(1行目)を太字＋背景色
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# 列幅を自動調整（おおよそ）
列幅 = {"A": 12, "B": 22, "C": 8, "D": 8, "E": 18}
for 列, 幅 in 列幅.items():
    ws.column_dimensions[列].width = 幅

# 保存
wb.save("food_prices.xlsx")
print("food_prices.xlsx を作成しました！")