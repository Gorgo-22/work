import anthropic
import os
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

# Excelファイルから食材データを読み込む
wb = load_workbook("food_prices.xlsx")
ws = wb["食材価格"]

food_items = []
# 2行目から最終行まで読み込む（1行目はヘッダー）
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:  # 空行はスキップ
        continue
    food_items.append({
        "カテゴリ": row[0],
        "名前": row[1],
        "先月": int(row[2]),
        "今月": int(row[3]),
        "単位": row[4],
    })

print(f"=== {len(food_items)}品目のデータを読み込みました ===\n")

# 個別の価格テキストを作成
price_text = ""
for item in food_items:
    変化 = item["今月"] - item["先月"]
    方向 = "値上がり" if 変化 > 0 else "値下がり"
    price_text += f"・[{item['カテゴリ']}] {item['名前']}：{item['先月']}{item['単位']} → {item['今月']}{item['単位']}（{方向} {abs(変化)}）\n"

# カテゴリ別の集計を計算
カテゴリ集計 = {}
for item in food_items:
    cat = item["カテゴリ"]
    if cat not in カテゴリ集計:
        カテゴリ集計[cat] = {"先月合計": 0, "今月合計": 0, "件数": 0}
    カテゴリ集計[cat]["先月合計"] += item["先月"]
    カテゴリ集計[cat]["今月合計"] += item["今月"]
    カテゴリ集計[cat]["件数"] += 1

# カテゴリ集計テキストを作成
集計_text = ""
for cat, 値 in カテゴリ集計.items():
    先月平均 = 値["先月合計"] / 値["件数"]
    今月平均 = 値["今月合計"] / 値["件数"]
    変動率 = (今月平均 - 先月平均) / 先月平均 * 100
    集計_text += f"・{cat}（{値['件数']}品目）：平均価格 {先月平均:.0f} → {今月平均:.0f}（{変動率:+.1f}%）\n"

prompt = f"""
以下の食材価格の変動について、経営会議向けに簡潔なコメントを200字以内で作成してください。

【個別品目の変動】
{price_text}

【カテゴリ別の平均価格変動】
{集計_text}

コメントの冒頭で全体傾向を述べた後、特に変動の大きいカテゴリと品目を取り上げてください。
コメントは「今月の食材価格は」で始めてください。
"""

message = client.messages.create(
    model="claude-opus-4-5",
    max_tokens=1024,
    messages=[{"role": "user", "content": prompt}]
)

print("=== カテゴリ別集計 ===")
print(集計_text)
print("=== 経営会議向けコメント ===")
print(message.content[0].text)